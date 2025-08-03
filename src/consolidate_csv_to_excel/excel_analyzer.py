import json
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill

from src.consolidate_csv_to_excel.custom_logger import CustomLogger

_HEADER_ROW = 1
_DATA_START_ROW = _HEADER_ROW + 1
_ZERO_BASED_INDEX_OFFSET = 1
_PROCESSING_TIME_COLUMN = 3 - _ZERO_BASED_INDEX_OFFSET
_ALERT_DETAIL_COLUMN = 4 - _ZERO_BASED_INDEX_OFFSET

_TRANSPARENT = "FF"
_YELLOW = "FFFF7F"
_GRAY = "7F7F7F"
_YELLOW_WITH_TRANSPARENT = _TRANSPARENT + _YELLOW
_GRAY_WITH_TRANSPARENT = _TRANSPARENT + _GRAY


class ExcelAnalyzer:
    _logger = CustomLogger.get_logger()

    def __init__(self, workbook: Workbook) -> None:
        self._workbook = workbook
        self._threshold_exceeded_sheets: set[str] = set()
        self._anomaly_detected_sheets: set[str] = set()

    @staticmethod
    def _highlight_cell(cell: Cell, color_code: str) -> None:
        pattern_fill = PatternFill(start_color=color_code, fill_type="solid")
        cell.fill = pattern_fill

    @staticmethod
    def _calculate_color_based_on_excess_ratio(
        processing_time: int, threshold: int
    ) -> str:
        excess_ratio = (processing_time - threshold) / threshold
        clamped_excess_ratio = min(excess_ratio, 1)

        MAX_GREEN_VALUE = 255
        MIN_GREEN_VALUE = MAX_GREEN_VALUE / 2
        green_value = int(
            MAX_GREEN_VALUE
            - (MAX_GREEN_VALUE - MIN_GREEN_VALUE) * clamped_excess_ratio
        )

        green_hex_value = f"{green_value:02X}"
        color_code = f"FF{green_hex_value}7F"

        return color_code

    def _check_and_highlight_processing_time(
        self,
        processing_time_cell: Cell,
        threshold: int,
    ) -> bool:
        processing_time_value = processing_time_cell.value

        if processing_time_value:
            try:
                processing_time_seconds = int(
                    processing_time_value.rstrip("s")
                )

                if processing_time_seconds >= threshold:
                    color_code = self._calculate_color_based_on_excess_ratio(
                        processing_time_seconds, threshold
                    )
                    self._highlight_cell(processing_time_cell, color_code)
                    return True
            except ValueError:
                self._logger.warning(
                    f"Invalid processing time value: {processing_time_value}"
                )

        return False

    def _check_and_highlight_alert_detail(
        self,
        alert_detail_cell: Cell,
    ) -> bool:
        alert_detail_value = alert_detail_cell.value

        if alert_detail_value:
            try:
                alert_detail_data = json.loads(alert_detail_value)
                if any(
                    item.get("random_key") is True
                    for item in alert_detail_data
                ):
                    self._highlight_cell(
                        alert_detail_cell, _YELLOW_WITH_TRANSPARENT
                    )
                    return True
            except json.JSONDecodeError:
                self._logger.warning(
                    f"Invalid JSON format found: {alert_detail_value}"
                )
        return False

    def _log_detected_anomalies(self, sheet_name: str) -> None:
        if sheet_name in self._threshold_exceeded_sheets:
            self._logger.warning(
                f"Processing time threshold exceeded: {sheet_name}"
            )

        if sheet_name in self._anomaly_detected_sheets:
            self._logger.warning(f"Anomaly value detected: {sheet_name}")

    def highlight_cells_and_sheet_tabs_by_criteria(
        self, threshold: int
    ) -> None:
        self._logger.info("Starting to highlight.")
        total_sheets = len(self._workbook.sheetnames)

        for current_sheet_number, sheet_name in enumerate(
            self._workbook.sheetnames, start=1
        ):
            sheet = self._workbook[sheet_name]
            has_highlighted_cell = False

            for row in sheet.iter_rows(min_row=_DATA_START_ROW):
                processing_time_cell = row[_PROCESSING_TIME_COLUMN]
                alert_detail_cell = row[_ALERT_DETAIL_COLUMN]

                if self._check_and_highlight_processing_time(
                    processing_time_cell, threshold
                ):
                    self._threshold_exceeded_sheets.add(sheet_name)
                    has_highlighted_cell = True

                if self._check_and_highlight_alert_detail(alert_detail_cell):
                    self._anomaly_detected_sheets.add(sheet_name)
                    has_highlighted_cell = True

            if has_highlighted_cell:
                sheet.sheet_properties.tabColor = _YELLOW_WITH_TRANSPARENT
                self._log_detected_anomalies(sheet_name)

            self._logger.info(
                f"Analyzed sheet: {sheet_name}."
                f" ({current_sheet_number}/{total_sheets})"
            )
        self._logger.info("Highlighting completed.")

    def _create_new_order(self) -> List[str]:
        yellow_sheets = []
        gray_sheets = []
        other_sheets = []

        for sheet_name in self._workbook.sheetnames:
            sheet_tab_color = self._workbook[
                sheet_name
            ].sheet_properties.tabColor

            if sheet_tab_color is None:
                other_sheets.append(sheet_name)
            else:
                sheet_color_value = sheet_tab_color.value

                if sheet_color_value == _YELLOW_WITH_TRANSPARENT:
                    yellow_sheets.append(sheet_name)
                elif sheet_color_value == _GRAY_WITH_TRANSPARENT:
                    gray_sheets.append(sheet_name)

        return yellow_sheets + other_sheets + gray_sheets

    def reorder_sheets_by_color(self) -> None:
        self._logger.info("Starting to reorder.")
        new_order = self._create_new_order()

        total_sheets = len(self._workbook.sheetnames)
        for current_sheet_number, sheet_name in enumerate(new_order, start=1):
            self._workbook.move_sheet(sheet_name, total_sheets)
            self._logger.info(
                f"Reordered sheet: {sheet_name}."
                f" ({current_sheet_number}/{total_sheets})"
            )
        self._logger.info("Reordering completed.")

    def get_analysis_results(self) -> Dict[str, set[str]]:
        return {
            "threshold_exceeded": self._threshold_exceeded_sheets,
            "anomaly_detected": self._anomaly_detected_sheets,
        }
