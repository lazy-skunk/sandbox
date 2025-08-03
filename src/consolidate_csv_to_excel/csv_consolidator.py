from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import Workbook

from src.consolidate_csv_to_excel.custom_logger import CustomLogger

_TRANSPARENT = "FF"
_GRAY = "7F7F7F"
_GRAY_WITH_TRANSPARENT = _TRANSPARENT + _GRAY


class CSVConsolidator:
    _logger = CustomLogger.get_logger()

    def __init__(self, writer: pd.ExcelWriter, workbook: Workbook) -> None:
        self._writer = writer
        self._workbook = workbook
        self._merge_failed_info: set[str] = set()

    def _create_sentinel_sheet(self) -> None:
        pd.DataFrame({"A": ["SENTINEL_SHEET"]}).to_excel(
            self._writer,
            sheet_name="SENTINEL_SHEET",
            index=False,
            header=False,
        )

    def _create_sheet_from_csv(self, sheet_name: str, csv_path: str) -> None:
        try:
            df = pd.read_csv(csv_path)
            df.to_excel(self._writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            self._logger.error(f"Failed to read CSV file at {csv_path}: {e}")
            self._merge_failed_info.add(sheet_name)

    def _create_no_csv_sheet(self, sheet_name: str) -> None:
        df_for_no_csv = pd.DataFrame({"A": ["No CSV file found."]})
        df_for_no_csv.to_excel(
            self._writer, sheet_name=sheet_name, index=False, header=False
        )

        self._writer.sheets[sheet_name].sheet_properties.tabColor = (
            _GRAY_WITH_TRANSPARENT
        )

    def _create_sheets(
        self, csv_paths_for_each_date: dict[str, Path | None]
    ) -> None:
        total_targets = len(csv_paths_for_each_date)

        for current_target_number, (sheet_name, csv_path) in enumerate(
            csv_paths_for_each_date.items(), start=1
        ):
            if csv_path:
                self._create_sheet_from_csv(sheet_name, csv_path)
            else:
                self._create_no_csv_sheet(sheet_name)

            self._logger.info(
                f"Added sheet: {sheet_name}."
                f" ({current_target_number}/{total_targets})"
            )

    def _delete_sentinel_sheet(self) -> None:
        if "SENTINEL_SHEET" in self._workbook.sheetnames:
            del self._workbook["SENTINEL_SHEET"]

    def consolidate_csvs_to_excel(
        self, csv_paths_for_each_date: dict[str, Path | None]
    ) -> None:
        self._logger.info("Starting to merge.")

        self._create_sentinel_sheet()
        self._create_sheets(csv_paths_for_each_date)
        self._delete_sentinel_sheet()

        self._logger.info("Merging completed.")

    def get_merge_failed_info(self) -> Dict[str, set[str]]:
        return {"merge_failed": self._merge_failed_info}
